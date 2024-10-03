import openpyxl
from openpyxl.styles import Alignment, numbers
from openpyxl.utils import get_column_letter
import csv
from sales_history_viewer import SalesHistoryViewer
from datetime import datetime
import sys
import requests
import pickle
import os
from PyQt5.QtWidgets import QFileDialog, QDialog, QLabel, QApplication, QTreeWidget, QTreeWidgetItem, QShortcut, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QCheckBox, QLabel, QMessageBox, QMainWindow, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QInputDialog
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtGui import QKeySequence, QFont

def is_valid_password(password):
    if ' ' in password or any(ord(char) > 127 for char in password):
        return False
    return True

class LoginForm(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.load_user_data()

    def initUI(self):
        self.setWindowTitle('Giriş Formu')
        self.setGeometry(300, 300, 300, 200)

        layout = QVBoxLayout()

        self.username = QLineEdit(self)
        self.username.setPlaceholderText('Kullanıcı adı')
        layout.addWidget(self.username)

        self.password = QLineEdit(self)
        self.password.setPlaceholderText('Şifre')
        self.password.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password)

        hbox_checkbox = QHBoxLayout()
        self.remember = QCheckBox('Beni Hatırla', self)
        self.show_password = QCheckBox('Şifreyi göster', self)
        self.show_password.stateChanged.connect(self.toggle_password_visibility)
        hbox_checkbox.addWidget(self.remember)
        hbox_checkbox.addWidget(self.show_password)
        layout.addLayout(hbox_checkbox)

        hbox_buttons = QHBoxLayout()
        self.login_button = QPushButton('Giriş yap', self)
        self.register_button = QPushButton('Kayıt ol', self)
        hbox_buttons.addWidget(self.login_button)
        hbox_buttons.addWidget(self.register_button)
        layout.addLayout(hbox_buttons)

        self.setLayout(layout)

        self.register_button.clicked.connect(self.open_register_form)
        self.login_button.clicked.connect(self.login)

    def toggle_password_visibility(self, state):
        if state == 2:
            self.password.setEchoMode(QLineEdit.Normal)
        else:
            self.password.setEchoMode(QLineEdit.Password)

    def open_register_form(self):
        self.register_form = RegisterForm()
        self.register_form.show()

    def login(self):
        username = self.username.text()
        password = self.password.text()
        
        if not username or not password:
            QMessageBox.warning(self, 'Hata', 'Lütfen kullanıcı adı ve şifre girin.')
            return

        if not is_valid_password(password):
            QMessageBox.warning(self, 'Hata', 'Şifre boşluk veya Türkçe karakter içermemelidir.')
            return
        
        response = requests.post('http://localhost/login.php', data={
            'username': username,
            'password': password
        })

        if response.text == 'success':
            QMessageBox.information(self, 'Başarılı', 'Giriş başarılı!')
            if self.remember.isChecked():
                self.save_user_data(username, password)
            self.open_main_menu()
        else:
            QMessageBox.warning(self, 'Hata', 'Giriş başarısız. Kullanıcı adı veya şifre hatalı.')

    def save_user_data(self, username, password):
        data = {'username': username, 'password': password}
        with open('data/user.pkl', 'wb') as f:
            pickle.dump(data, f)

    def load_user_data(self):
        if os.path.exists('data/user.pkl'):
            with open('data/user.pkl', 'rb') as f:
                data = pickle.load(f)
                self.username.setText(data['username'])
                self.password.setText(data['password'])
                self.remember.setChecked(True)

    def open_main_menu(self):
        self.main_menu = MainMenu()
        self.main_menu.show()
        self.close()


class RegisterForm(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Kayıt Formu')
        self.setGeometry(400, 300, 300, 300)

        layout = QVBoxLayout()

        self.fullname = QLineEdit(self)
        self.fullname.setPlaceholderText('Ad Soyad')
        layout.addWidget(self.fullname)

        self.username = QLineEdit(self)
        self.username.setPlaceholderText('Kullanıcı Adı')
        layout.addWidget(self.username)

        self.password = QLineEdit(self)
        self.password.setPlaceholderText('Şifre')
        self.password.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password)

        self.password_confirm = QLineEdit(self)
        self.password_confirm.setPlaceholderText('Şifre tekrar')
        self.password_confirm.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_confirm)

        self.contact = QLineEdit(self)
        self.contact.setPlaceholderText('İletişim numarası')
        layout.addWidget(self.contact)

        self.show_password = QCheckBox('Şifreyi göster', self)
        self.show_password.stateChanged.connect(self.toggle_password_visibility)
        layout.addWidget(self.show_password)

        self.register_button = QPushButton('Kayıt ol', self)
        layout.addWidget(self.register_button)

        self.setLayout(layout)

        self.register_button.clicked.connect(self.register)

    def toggle_password_visibility(self, state):
        if state == 2:
            self.password.setEchoMode(QLineEdit.Normal)
            self.password_confirm.setEchoMode(QLineEdit.Normal)
        else:
            self.password.setEchoMode(QLineEdit.Password)
            self.password_confirm.setEchoMode(QLineEdit.Password)

    def register(self):
        if not all([self.fullname.text(), self.username.text(), self.password.text(), 
                    self.password_confirm.text(), self.contact.text()]):
            QMessageBox.warning(self, 'Hata', 'Lütfen tüm alanları doldurun.')
        elif self.password.text() != self.password_confirm.text():
            QMessageBox.warning(self, 'Hata', 'Şifreler eşleşmiyor.')
        elif not is_valid_password(self.password.text()):
            QMessageBox.warning(self, 'Hata', 'Şifre boşluk veya Türkçe karakter içermemelidir.')
        else:
            response = requests.post('http://localhost/register.php', data={
                'fullname': self.fullname.text(),
                'username': self.username.text(),
                'password': self.password.text(),
                'contact': self.contact.text()
            })

            if response.text == 'success':
                QMessageBox.information(self, 'Başarılı', 'Kayıt başarıyla tamamlandı.')
                self.close()
            else:
                QMessageBox.warning(self, 'Hata', 'Kayıt işlemi başarısız oldu. Lütfen tekrar deneyin.')

class MainMenu(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Ana Menü')
        self.setGeometry(300, 300, 300, 250)

        layout = QVBoxLayout()

        self.barcode_button = QPushButton('Barkod Okuyucu', self)
        self.barcode_button.clicked.connect(self.open_barcode_reader)
        layout.addWidget(self.barcode_button)

        self.ledger_button = QPushButton('Veresiye Defteri', self)
        layout.addWidget(self.ledger_button)

        self.product_list_button = QPushButton('Ürün Listesi', self)
        self.product_list_button.clicked.connect(self.open_product_list)
        layout.addWidget(self.product_list_button)

        self.setLayout(layout)

    def open_barcode_reader(self):
        self.barcode_reader = BarcodeReader()
        self.barcode_reader.show()
        self.close()

    def open_product_list(self):
        self.product_list = ProductList()
        self.product_list.show()
        self.close()
class ProductList(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.load_products()

    def initUI(self):
        self.setWindowTitle('Ürün Listesi')
        self.showFullScreen()

        layout = QVBoxLayout()

        # Arama alanı
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit(self)
        self.search_input.setPlaceholderText('Ürün ara...')
        self.search_input.textChanged.connect(self.search_products)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)

        # Ürün tablosu
        self.product_table = QTableWidget(self)
        self.product_table.setColumnCount(5)
        self.product_table.setHorizontalHeaderLabels(['Barkod', 'Ürün Adı', 'Alış Fiyatı', 'Satış Fiyatı', ''])
        self.product_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.product_table.setEditTriggers(QAbstractItemView.DoubleClicked)
        
        # Başlık satırının arka plan rengini gri yap
        header = self.product_table.horizontalHeader()
        header.setStyleSheet("QHeaderView::section { background-color: lightgray }")
        
        layout.addWidget(self.product_table)

        # Butonlar
        button_layout = QHBoxLayout()
        
        self.save_button = QPushButton('Değişiklikleri Kaydet', self)
        self.save_button.clicked.connect(self.save_changes)
        button_layout.addWidget(self.save_button)

        self.export_button = QPushButton('Ürünleri Dışarı Aktar', self)
        self.export_button.clicked.connect(self.export_products)
        button_layout.addWidget(self.export_button)

        self.import_button = QPushButton('Ürünleri İçe Aktar', self)
        self.import_button.clicked.connect(self.import_products)
        button_layout.addWidget(self.import_button)

        self.back_button = QPushButton('Ana Menüye Dön', self)
        self.back_button.clicked.connect(self.go_back)
        button_layout.addWidget(self.back_button)

        layout.addLayout(button_layout)

        self.setLayout(layout)

    def load_products(self):
        if os.path.exists('data/products.pkl'):
            with open('data/products.pkl', 'rb') as f:
                self.products = pickle.load(f)
        else:
            self.products = {}
        
        self.update_table()

    def update_table(self, filtered_products=None):
        self.product_table.setRowCount(0)
        products_to_display = filtered_products if filtered_products is not None else self.products
        
        for barcode, product in products_to_display.items():
            row_position = self.product_table.rowCount()
            self.product_table.insertRow(row_position)
            
            self.product_table.setItem(row_position, 0, QTableWidgetItem(barcode))
            self.product_table.setItem(row_position, 1, QTableWidgetItem(product['name']))
            self.product_table.setItem(row_position, 2, QTableWidgetItem(str(product['buy_price'])))
            self.product_table.setItem(row_position, 3, QTableWidgetItem(str(product['sell_price'])))
            
            delete_button = QPushButton('Sil')
            delete_button.clicked.connect(lambda _, b=barcode: self.delete_product(b))
            self.product_table.setCellWidget(row_position, 4, delete_button)

    def save_changes(self):
        for row in range(self.product_table.rowCount()):
            barcode = self.product_table.item(row, 0).text()
            name = self.product_table.item(row, 1).text()
            buy_price = float(self.product_table.item(row, 2).text())
            sell_price = float(self.product_table.item(row, 3).text())
            
            self.products[barcode] = {
                'name': name,
                'buy_price': buy_price,
                'sell_price': sell_price
            }
        
        with open('data/products.pkl', 'wb') as f:
            pickle.dump(self.products, f)
        
        QMessageBox.information(self, 'Başarılı', 'Değişiklikler kaydedildi!')

    def delete_product(self, barcode):
        reply = QMessageBox.question(self, 'Ürünü Sil', f'"{self.products[barcode]["name"]}" ürününü silmek istediğinizden emin misiniz?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            del self.products[barcode]
            self.update_table()

    def go_back(self):
        self.main_menu = MainMenu()
        self.main_menu.show()
        self.close()

    def search_products(self):
        search_text = self.search_input.text().lower()
        if not search_text:
            self.update_table()
            return

        filtered_products = {}
        for barcode, product in self.products.items():
            if (search_text in barcode.lower() or
                search_text in product['name'].lower() or
                search_text in str(product['buy_price']).lower() or
                search_text in str(product['sell_price']).lower()):
                filtered_products[barcode] = product

        self.update_table(filtered_products)

    def export_products(self):
        export_dialog = ExportDialog(self)
        if export_dialog.exec_():
            export_type = export_dialog.selected_type
            default_path = os.path.expanduser("~/Desktop")
            if export_type == 'CSV':
                file_path, _ = QFileDialog.getSaveFileName(self, "CSV Dosyasını Kaydet", default_path, "CSV Files (*.csv)")
                if file_path:
                    self.export_to_csv(file_path)
            elif export_type == 'EXCEL':
                file_path, _ = QFileDialog.getSaveFileName(self, "Excel Dosyasını Kaydet", default_path, "Excel Files (*.xlsx)")
                if file_path:
                    self.export_to_excel(file_path)

    def export_to_csv(self, file_path):
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Barkod', 'Ürün Adı', 'Alış Fiyatı', 'Satış Fiyatı'])
            for barcode, product in self.products.items():
                writer.writerow([str(barcode), product['name'], product['buy_price'], product['sell_price']])
        QMessageBox.information(self, 'Başarılı', f'Ürünler {file_path} dosyasına aktarıldı.')

    def export_to_excel(self, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['Barkod', 'Ürün Adı', 'Alış Fiyatı', 'Satış Fiyatı']
        ws.append(headers)
        
        for barcode, product in self.products.items():
            ws.append([str(barcode), product['name'], product['buy_price'], product['sell_price']])

        # Hücreleri ortala ve sütun genişliklerini ayarla
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Sütun harfini al
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if column == 'A':  # Barkod sütunu
                    cell.number_format = '@'  # Metin formatı
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        QMessageBox.information(self, 'Başarılı', f'Ürünler {file_path} dosyasına aktarıldı.')

    def update_or_add_product(self, barcode, product_data):
        if barcode in self.products:
            # Eğer ürün zaten varsa, önce onu sil
            del self.products[barcode]
        
        # Yeni ürünü ekle
        self.products[barcode] = {
            'name': product_data['name'],
            'buy_price': float(product_data['buy_price']),
            'sell_price': float(product_data['sell_price'])
        }

    def import_from_csv(self, file_path):
        with open(file_path, 'r', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                barcode = row['Barkod']
                self.update_or_add_product(barcode, row)
        self.save_changes_for_doc()
        self.update_table()
        QMessageBox.information(self, 'Başarılı', f'Ürünler {file_path} dosyasından içe aktarıldı.')

    def import_from_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            barcode = str(row[headers.index('Barkod')])
            product_data = {
                'name': row[headers.index('Ürün Adı')],
                'buy_price': float(row[headers.index('Alış Fiyatı')]),
                'sell_price': float(row[headers.index('Satış Fiyatı')])
            }
            self.update_or_add_product(barcode, product_data)
        self.save_changes_for_doc()
        self.update_table()
        QMessageBox.information(self, 'Başarılı', f'Ürünler {file_path} dosyasından içe aktarıldı.')

    def save_changes_for_doc(self):
        with open('data/products.pkl', 'wb') as f:
            pickle.dump(self.products, f)
        
    def update_table(self):
        self.product_table.setRowCount(0)
        for barcode, product in self.products.items():
            row_position = self.product_table.rowCount()
            self.product_table.insertRow(row_position)
            
            self.product_table.setItem(row_position, 0, QTableWidgetItem(barcode))
            self.product_table.setItem(row_position, 1, QTableWidgetItem(product['name']))
            self.product_table.setItem(row_position, 2, QTableWidgetItem(str(product['buy_price'])))
            self.product_table.setItem(row_position, 3, QTableWidgetItem(str(product['sell_price'])))
            
            delete_button = QPushButton('Sil')
            delete_button.clicked.connect(lambda _, b=barcode: self.delete_product(b))
            self.product_table.setCellWidget(row_position, 4, delete_button)



    def import_products(self):
        default_path = os.path.expanduser("~/Desktop")
        file_path, _ = QFileDialog.getOpenFileName(self, "Dosya Seç", default_path, "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if file_path:
            if file_path.endswith('.csv'):
                self.import_from_csv(file_path)
            elif file_path.endswith('.xlsx'):
                self.import_from_excel(file_path)

class ExportDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Dışarı Aktarma Türü')
        self.setGeometry(300, 300, 250, 100)

        layout = QVBoxLayout()
        
        label = QLabel('Dışarı aktarma türü:')
        layout.addWidget(label)

        self.csv_button = QPushButton('CSV')
        self.csv_button.clicked.connect(lambda: self.select_type('CSV'))
        layout.addWidget(self.csv_button)

        self.excel_button = QPushButton('EXCEL')
        self.excel_button.clicked.connect(lambda: self.select_type('EXCEL'))
        layout.addWidget(self.excel_button)

        self.setLayout(layout)
        self.selected_type = None

    def select_type(self, export_type):
        self.selected_type = export_type
        self.accept()



class BarcodeReader(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.cart = {}
        self.load_products()
        self.load_order_history()


    def initUI(self):
        self.setWindowTitle('Barkod Okuyucu')
        self.showFullScreen()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        self.barcode_input = QLineEdit(self)
        self.barcode_input.setPlaceholderText('Barkod girin')
        self.barcode_input.returnPressed.connect(self.add_to_cart)
        layout.addWidget(self.barcode_input)

        button_layout = QHBoxLayout()
        self.sell_button = QPushButton('Satış Yap (F1)', self)
        self.remove_button = QPushButton('Ürün Sil (F8)', self)
        self.clear_button = QPushButton('Sepeti Boşalt (F9)', self)
        self.change_quantity_button = QPushButton('Adet Seç (F5)', self)

        self.view_history_button = QPushButton('Satış Geçmişini Görüntüle', self)
        self.view_history_button.clicked.connect(self.open_sales_history)
        self.view_history_button.setMinimumHeight(50)
        layout.addWidget(self.view_history_button)

        button_height = 50
        for button in [self.sell_button, self.remove_button, self.clear_button, self.change_quantity_button]:
            button.setMinimumHeight(button_height)

        button_layout.addWidget(self.sell_button)
        button_layout.addWidget(self.remove_button)
        button_layout.addWidget(self.clear_button)
        button_layout.addWidget(self.change_quantity_button)
        layout.addLayout(button_layout)

        self.cart_table = QTableWidget(self)
        self.cart_table.setColumnCount(4)
        self.cart_table.setHorizontalHeaderLabels(['Ürün Barkodu', 'Ürün Adı', 'Fiyatı', 'Adet'])
        self.cart_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.cart_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.cart_table)

        self.total_label = QLabel('Toplam: 0.00 TL', self)
        self.total_label.setAlignment(Qt.AlignLeft)
        font = QFont()
        font.setPointSize(16)
        self.total_label.setFont(font)
        layout.addWidget(self.total_label)

        self.add_product_button = QPushButton('Ürün Ekle', self)
        self.add_product_button.clicked.connect(self.open_add_product)
        layout.addWidget(self.add_product_button, alignment=Qt.AlignRight)

        self.sell_button.clicked.connect(self.sell)
        self.remove_button.clicked.connect(self.remove_item)
        self.clear_button.clicked.connect(self.clear_cart)
        self.change_quantity_button.clicked.connect(self.change_quantity)

        QShortcut(QKeySequence("F1"), self, self.sell)
        QShortcut(QKeySequence("F8"), self, self.remove_item)
        QShortcut(QKeySequence("F9"), self, self.clear_cart)
        QShortcut(QKeySequence("F5"), self, self.change_quantity)

        # Otomatik odaklanma
        self.barcode_input.setFocus()  # Burada odaklanmayı ayarlıyoruz.

    
    def add_to_cart(self):
        barcode = self.barcode_input.text()
        if barcode in self.products:
            product = self.products[barcode]
            if barcode in self.cart:
                self.cart[barcode]['quantity'] += 1
                self.update_cart_table()
            else:
                self.cart[barcode] = {
                    'name': product['name'],
                    'price': product['sell_price'],
                    'quantity': 1
                }
                self.add_to_cart_table(barcode)
            self.update_total()
        else:
            QMessageBox.warning(self, 'Hata', 'Ürün bulunamadı!')
        self.barcode_input.clear()

    def add_to_cart_table(self, barcode):
        row_position = self.cart_table.rowCount()
        self.cart_table.insertRow(row_position)
        self.cart_table.setItem(row_position, 0, QTableWidgetItem(barcode))
        self.cart_table.setItem(row_position, 1, QTableWidgetItem(self.cart[barcode]['name']))
        self.cart_table.setItem(row_position, 2, QTableWidgetItem(str(self.cart[barcode]['price'])))
        self.cart_table.setItem(row_position, 3, QTableWidgetItem(str(self.cart[barcode]['quantity'])))

    def update_cart_table(self):
        self.cart_table.setRowCount(0)
        for barcode, item in self.cart.items():
            row_position = self.cart_table.rowCount()
            self.cart_table.insertRow(row_position)
            self.cart_table.setItem(row_position, 0, QTableWidgetItem(barcode))
            self.cart_table.setItem(row_position, 1, QTableWidgetItem(item['name']))
            self.cart_table.setItem(row_position, 2, QTableWidgetItem(str(item['price'])))
            self.cart_table.setItem(row_position, 3, QTableWidgetItem(str(item['quantity'])))

    def update_total(self):
        total = sum(item['price'] * item['quantity'] for item in self.cart.values())
        self.total_label.setText(f'Toplam: {total:.2f} TL')

    def sell(self):
        if not self.cart:
            QMessageBox.warning(self, 'Hata', 'Sepet boş!')
            return
        
        total = sum(item['price'] * item['quantity'] for item in self.cart.values())
        sale_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        sale_record = {
            'time': sale_time,
            'total': total,
            'items': [
                {
                    'barcode': barcode,
                    'name': item['name'],
                    'price': item['price'],
                    'quantity': item['quantity']
                } for barcode, item in self.cart.items()
            ]
        }
        
        self.order_history.append(sale_record)
        self.save_order_history()
        
        QMessageBox.information(self, 'Başarılı', 'Satış tamamlandı!')
        self.clear_cart()

    def load_order_history(self):
        if os.path.exists('data/orderHistory.pkl'):
            with open('data/orderHistory.pkl', 'rb') as f:
                self.order_history = pickle.load(f)
        else:
            self.order_history = []

    def save_order_history(self):
        with open('data/orderHistory.pkl', 'wb') as f:
            pickle.dump(self.order_history, f)

    def open_sales_history(self):
        self.sales_history_viewer = SalesHistoryViewer(self.order_history)
        self.sales_history_viewer.show()

    def remove_item(self):
        current_row = self.cart_table.currentRow()
        if current_row >= 0:
            barcode = self.cart_table.item(current_row, 0).text()
            del self.cart[barcode]
            self.update_cart_table()
            self.update_total()

    def clear_cart(self):
        self.cart.clear()
        self.cart_table.setRowCount(0)
        self.update_total()

    def change_quantity(self):
        current_row = self.cart_table.currentRow()
        if current_row >= 0:
            barcode = self.cart_table.item(current_row, 0).text()
            quantity, ok = QInputDialog.getInt(self, "Adet Seç", "Yeni adet:", 
                                               self.cart[barcode]['quantity'], 1, 100)
            if ok:
                self.cart[barcode]['quantity'] = quantity
                self.update_cart_table()
                self.update_total()

    def load_products(self):
        if os.path.exists('data/products.pkl'):
            with open('data/products.pkl', 'rb') as f:
                self.products = pickle.load(f)
        else:
            self.products = {}

    def save_products(self):
        with open('data/products.pkl', 'wb') as f:
            pickle.dump(self.products, f)

    def open_add_product(self):
        self.add_product_form = AddProductForm(self)
        self.add_product_form.show()



class AddProductForm(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Ürün Ekle')
        self.setGeometry(400, 300, 300, 200)

        layout = QVBoxLayout()

        self.barcode_input = QLineEdit(self)
        self.barcode_input.setPlaceholderText('Barkod')
        layout.addWidget(self.barcode_input)

        self.name_input = QLineEdit(self)
        self.name_input.setPlaceholderText('Ürün Adı')
        layout.addWidget(self.name_input)

        self.buy_price_input = QLineEdit(self)
        self.buy_price_input.setPlaceholderText('Alış Fiyatı')
        layout.addWidget(self.buy_price_input)

        self.sell_price_input = QLineEdit(self)
        self.sell_price_input.setPlaceholderText('Satış Fiyatı')
        layout.addWidget(self.sell_price_input)

        self.save_button = QPushButton('Kaydet', self)
        self.save_button.clicked.connect(self.save_product)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

    def save_product(self):
        barcode = self.barcode_input.text()
        name = self.name_input.text()
        buy_price = float(self.buy_price_input.text())
        sell_price = float(self.sell_price_input.text())

        if not all([barcode, name, buy_price, sell_price]):
            QMessageBox.warning(self, 'Hata', 'Lütfen tüm alanları doldurun!')
            return

        self.parent.products[barcode] = {
            'name': name,
            'buy_price': buy_price,
            'sell_price': sell_price
        }
        self.parent.save_products()
        QMessageBox.information(self, 'Başarılı', 'Ürün kaydedildi!')
        self.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    login_form = LoginForm()
    login_form.show()
    sys.exit(app.exec_())